from openpyxl import Workbook
from openpyxl import load_workbook
# wb2 = load_workbook('Test.xlsx')
# wb = Workbook()
# ws = wb2.active

def create_sheet(title, sheetList):
    wb.create_sheet("{x}.xlsx".format(x = title))
    wb.save("{x}.xlsx".format(x = title))
    for name in sheetList:
        wb.create_sheet("{}".format(name))
        wb.save("{}".format(name))


# Word_Page = wb2['Sheet']
# Word = Word_Page['A']
# Definition = Word_Page['B']
# Etymology = Word_Page['C']
# Synonyms = Word_Page['D']
