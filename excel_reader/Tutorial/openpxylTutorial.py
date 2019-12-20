from openpyxl import Workbook
wb = Workbook()

ws = wb.active #This is set to 0 by default. Unless you modify its value, you will always get the first worksheet by using this method

ws1 = wb.create_sheet("Mysheet")#insert at end (default)
ws2 = wb.create_sheet("Mysheet", 0)# insert at first position
# ws3 = wb.create_sheet("Mysheet", -1)# insert at the punultimate position

ws.title = "New Title"
ws.sheet_properties.tabColor = "1072BA"

ws3 = wb["New Title"]

target = wb.copy_worksheet(ws3)

c = ws3['A5']
ws3['A4'] = 4

c.value = 12


for row in ws3.values:
    for value in row:
        print(value)

wb.save('test.xlsx')

